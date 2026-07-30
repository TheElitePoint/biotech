"""Export the final company list to output/final.xlsx.

A single clean sheet - the companies that survived every gate, without the
run-tracking machinery of the weekly report. This is the shareable artefact.

    python save_final.py              # search pipeline output (review_queue.csv)
    python save_final.py --universe   # registry universe output (universe_intent.csv)
"""

from __future__ import annotations

import argparse
import csv
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "output"
SOURCE = OUT / "review_queue.csv"
HISTORY = OUT / "company_history.csv"
FINAL = OUT / "final.xlsx"

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PRIORITY_FILL = {
    "A": PatternFill("solid", fgColor="C6EFCE"),
    "B": PatternFill("solid", fgColor="D9EAD3"),
    "Review": PatternFill("solid", fgColor="FFF2CC"),
}

SEARCH_COLUMNS = [
    ("priority", "Priority", 9),
    ("score", "Score", 7),
    ("canonical_company", "Company", 28),
    ("domain", "Domain", 24),
    ("signal_types", "Intent Trigger", 16),
    ("top_signal_date", "Signal Date", 12),
    ("modality", "Modality", 18),
    ("asset_stage", "Asset Stage", 22),
    ("bottlenecks", "Bottleneck", 32),
    ("top_signal_title", "Evidence", 56),
    ("top_signal_url", "Source URL", 44),
    ("open_question", "Open Question", 40),
]

UNIVERSE_COLUMNS = [
    ("priority", "Priority", 9),
    ("score", "Score", 7),
    ("canonical_company", "Company", 32),
    ("trigger", "Trigger Source", 13),
    ("signal_type", "Intent Type", 13),
    ("signal_date", "Signal Date", 12),
    ("modality", "Modality", 26),
    ("phases", "Trial Phases", 20),
    ("statuses", "Trial Status", 26),
    ("trial_count", "Trials", 8),
    ("conditions", "Disease Areas", 34),
    ("interventions", "Interventions", 40),
    ("evidence", "Evidence", 46),
    ("evidence_url", "Source URL", 40),
    ("note", "Flag", 30),
]


def read(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))


def main() -> int:
    parser = argparse.ArgumentParser(description="Export the final company list")
    parser.add_argument(
        "--universe",
        action="store_true",
        help="Export the registry-universe list instead of the search pipeline output",
    )
    args = parser.parse_args()

    source = OUT / ("universe_intent.csv" if args.universe else "review_queue.csv")
    columns = UNIVERSE_COLUMNS if args.universe else SEARCH_COLUMNS

    rows = read(source)
    if not rows:
        cmd = (
            "python -m intent_pipeline.universe && python -m intent_pipeline.universe_intent"
            if args.universe
            else "python -m intent_pipeline.run"
        )
        raise SystemExit(f"{source} is empty. Run first:\n  {cmd}")

    # Drop anything the reviewer has already closed.
    closed = {
        h["company_key"]
        for h in read(HISTORY)
        if (h.get("reviewer_status") or "").strip().lower() in {"rejected", "do not contact"}
    }
    rows = [r for r in rows if r.get("company_key") not in closed]
    rows.sort(key=lambda r: -int(r.get("score") or 0))

    wb = Workbook()
    ws = wb.active
    ws.title = "Final List"
    ws.sheet_view.showGridLines = False

    ws.cell(row=1, column=1, value=f"Qualified Companies - {date.today().isoformat()}").font = Font(
        name=FONT, size=13, bold=True, color="1F3864"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))

    subtitle = (
        f"{len(rows)} companies passed the hard exclusions and scoring gates. "
        "Each still needs asset-ownership confirmation before outreach (SOP Gate 5)."
    )
    ws.cell(row=2, column=1, value=subtitle).font = Font(name=FONT, size=9, italic=True, color="595959")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))

    for col, (_, header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[4].height = 26

    for i, row in enumerate(rows, start=5):
        for col, (key, _, _) in enumerate(columns, start=1):
            value = row.get(key, "")
            if key == "score":
                value = int(value) if str(value).strip().isdigit() else value
            cell = ws.cell(row=i, column=col, value=value)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=key in ("bottlenecks", "top_signal_title", "open_question"),
            )
            if key in ("top_signal_url", "evidence_url") and value:
                cell.hyperlink = value
                cell.font = Font(name=FONT, size=10, color="0563C1", underline="single")
            elif key == "priority" and value in PRIORITY_FILL:
                cell.fill = PRIORITY_FILL[value]
                cell.alignment = Alignment(horizontal="center", vertical="top")
        ws.row_dimensions[i].height = 30

    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(columns))}{4 + len(rows)}"

    FINAL.parent.mkdir(parents=True, exist_ok=True)
    wb.save(FINAL)

    print(f"Saved {len(rows)} companies -> {FINAL}\n")

    from collections import Counter

    print(f"  priority: {dict(Counter(r['priority'] for r in rows))}")
    if args.universe:
        print(f"  trigger : {dict(Counter(r.get('trigger', '') for r in rows))}")
    print("\n  Top 12:")
    for r in rows[:12]:
        detail = r.get("trigger") or r.get("signal_types", "")
        print(f"    {r['priority']:<7} {r['score']:>3}  {r['canonical_company'][:34]:34} {detail}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
