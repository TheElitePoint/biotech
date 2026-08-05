"""Multi-source orchestrator: run every free source, score through the
existing SOP gates, and write exactly two result files.

    python -m intent_pipeline.orchestrate               # full run, all 8 sources
    python -m intent_pipeline.orchestrate --sources bing_news,pubmed
    python -m intent_pipeline.orchestrate --list-sources

Nothing here reimplements scoring or dedupe — every item flows through the
same signals.normalize/dedupe -> gates.route -> store.build_company_master
pipeline the Tavily-based intent_pipeline.run already uses. The only new
piece is `registry.py`, the seen-company ledger that stops an already-decided
company from re-entering the report, and the two-file writer below.
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import company_profile
from . import intents as intents_mod
from . import registry
from .gates import route
from .signals import Signal, canonical_name, dedupe, normalize
from .sources import company_site, company_size, jobs
from .sources.base import attach_known_company
from .suppression import contains as is_suppressed
from .suppression import load as load_suppression

RESULTS = Path(__file__).resolve().parent.parent / "results"
MASTER = RESULTS / "master.xlsx"
LATEST = RESULTS / "latest_run.xlsx"

# Per-source recency window in days. Bulk registries look back further than
# fast-moving news; biotech_rss ignores this (it's just "current feed").
DEFAULT_DAYS = {
    "clinicaltrials.gov": 180,
    "nih_reporter": 365,
    "europepmc": 365,
    "pubmed": 365,
    "bing_news": 365,
    "sec_edgar": 365,
    "crossref": 365,
    "biotech_rss": 30,
    "account_based": 180,
}

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name=FONT, size=10)
LINK_FONT = Font(name=FONT, size=10, color="0563C1", underline="single")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
PRIORITY_FILL = {
    "A": PatternFill("solid", fgColor="C6EFCE"),
    "B": PatternFill("solid", fgColor="D9EAD3"),
    "Review": PatternFill("solid", fgColor="FFF2CC"),
}
MOVEMENT_FILL = {"New": PatternFill("solid", fgColor="D9E1F2"), "Updated": PatternFill("solid", fgColor="FCE4D6")}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DECISIONS = '"Approved,Rejected,Watching,Needs Info"'

# No date columns anywhere in the results sheets, per standing instruction.
COLUMNS = [
    ("movement", "Movement", 11),
    ("priority", "Priority", 9),
    ("score", "Score", 7),
    ("canonical_company", "Company", 28),
    ("domain", "Domain", 22),
    ("employees", "Employees", 11),
    ("sources", "Sources", 26),
    ("signal_types", "Intent Trigger", 18),
    ("modality", "Modality", 20),
    ("bottlenecks", "Bottleneck", 30),
    ("top_signal_title", "Evidence Headline", 50),
    ("top_signal_url", "Evidence URL", 42),
    ("open_question", "Open Question", 36),
    ("times_seen", "Runs Seen", 10),
    ("reviewer_status", "DECISION", 14),
    ("reviewer", "Reviewer", 14),
    ("reviewer_notes", "Notes", 36),
]
REVIEWER_COLS = {"reviewer_status", "reviewer", "reviewer_notes"}


def run_sources(
    selected: list[str] | None,
    min_employees: int = 500,
    max_employees: int | None = None,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    names = selected or list(intents_mod.SOURCES)
    raw_items: list[dict[str, Any]] = []
    per_source_counts: dict[str, int] = {}

    for name in names:
        fetch = intents_mod.SOURCES.get(name)
        if fetch is None:
            print(f"  unknown source: {name}", file=sys.stderr)
            continue
        days = DEFAULT_DAYS.get(name, 365)
        try:
            if name == "biotech_rss":
                items = fetch()
            elif name == "account_based":
                items = fetch(days=days, min_employees=min_employees, max_employees=max_employees)
            else:
                items = fetch(days=days)
        except Exception as exc:  # noqa: BLE001 - one bad source must not kill the run
            print(f"  {name}: FAILED — {exc}")
            per_source_counts[name] = 0
            continue
        for item in items:
            item["_source_name"] = name
        raw_items.extend(items)
        per_source_counts[name] = len(items)
        print(f"  {name:<18} {len(items):4} raw items")

    return raw_items, per_source_counts


def build_signals(raw_items: list[dict[str, Any]]) -> list[Signal]:
    sigs = []
    for item in raw_items:
        sig = normalize(item)
        attach_known_company(sig, item)
        sig.source = item.get("_source_name", sig.source)
        sigs.append(sig)
    return dedupe(sigs)


def build_company_rows(records: list[tuple[Signal, Any]]) -> list[dict[str, Any]]:
    """Same collapse-to-one-row-per-company logic as store.build_company_master,
    plus a `sources` field (which free APIs corroborate this company) that the
    existing store module has no reason to know about.
    """
    suppressed = load_suppression()
    groups: dict[str, list[tuple[Signal, Any]]] = defaultdict(list)
    for sig, verdict in records:
        if verdict.decision == "Reject" and verdict.exclusion_reason:
            continue
        name = sig.company_candidate
        domain = sig.company_domain
        key = (domain or canonical_name(name or "").lower()).strip()
        if not key or is_suppressed(suppressed, key=key, name=name, domain=domain):
            continue
        groups[key].append((sig, verdict))

    rows = []
    for key, group in groups.items():
        group.sort(key=lambda g: g[1].score, reverse=True)
        best_sig, best_verdict = group[0]
        signal_types = sorted({s.signal_type for s, _ in group})
        sources = sorted({s.source for s, _ in group})
        bottlenecks = sorted({b for _, v in group for b in v.evidence.get("bottlenecks", [])})
        bonus = min(8, 4 * (len(signal_types) - 1))
        total = min(100, best_verdict.score + bonus)
        priority = "A" if total >= 80 else "B" if total >= 68 else "Review" if total >= 55 else "Reject"
        if best_verdict.decision != "Approve":
            priority = best_verdict.priority

        # Whether a *known-large* company gets floored to Review despite a low
        # score is decided later in main(), once every company's size is known —
        # see the "large account" floor there. That applies uniformly regardless
        # of which of the 9 sources surfaced the company, not just account_based.
        open_question = best_verdict.open_question

        rows.append(
            {
                "company_key": key,
                "canonical_company": best_sig.company_candidate or "",
                "domain": best_sig.company_domain or "",
                "priority": priority,
                "score": total,
                "signal_types": "; ".join(signal_types),
                "sources": "; ".join(sources),
                "modality": "; ".join(best_verdict.evidence.get("modality_approve", [])),
                "bottlenecks": "; ".join(bottlenecks),
                "top_signal_title": best_sig.title,
                "top_signal_url": best_sig.source_url,
                "open_question": open_question,
                "all_evidence_urls": " | ".join(s.source_url for s, _ in group[:6]),
            }
        )

    rows.sort(key=lambda r: (r["priority"] not in ("A", "B"), -r["score"]))
    return rows


def sync_decisions_from_master() -> int:
    """Read the reviewer's DECISION column out of last run's master.xlsx before
    it gets overwritten — same read-before-overwrite pattern as excel.py's
    sync_decisions(), pointed at the new results file instead.
    """
    if not MASTER.exists():
        return 0
    try:
        wb = load_workbook(MASTER, data_only=True)
    except Exception:
        return 0

    keys = [key for key, _, _ in COLUMNS]
    decisions: dict[str, dict[str, str]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            record = {k: ("" if v is None else str(v)) for k, v in zip(keys, row)}
            company = record.get("canonical_company", "").strip()
            domain = record.get("domain", "").strip()
            key = domain or company.lower()
            status = record.get("reviewer_status", "").strip()
            if key and status:
                decisions[key] = record
    return registry.record_decisions(decisions)


def _write_sheet(ws, rows: list[dict[str, Any]], title_note: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value=title_note).font = Font(name=FONT, size=11, bold=True, color="1F3864")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    for col, (_, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 26

    for i, row in enumerate(rows, start=4):
        for col, (key, _, _) in enumerate(COLUMNS, start=1):
            value = row.get(key, "")
            if key in ("score", "times_seen", "employees") and str(value).strip().isdigit():
                value = int(value)
            cell = ws.cell(row=i, column=col, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=key in ("bottlenecks", "top_signal_title", "open_question", "reviewer_notes"))
            if key == "top_signal_url" and value:
                cell.hyperlink = value
                cell.font = LINK_FONT
            elif key == "priority" and value in PRIORITY_FILL:
                cell.fill = PRIORITY_FILL[value]
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key == "movement" and value in MOVEMENT_FILL:
                cell.fill = MOVEMENT_FILL[value]
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key in REVIEWER_COLS:
                cell.fill = INPUT_FILL
        ws.row_dimensions[i].height = 28

    if rows:
        ws.freeze_panes = "D4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUMNS))}{3 + len(rows)}"
        decision_col = get_column_letter(next(i for i, (k, _, _) in enumerate(COLUMNS, start=1) if k == "reviewer_status"))
        dv = DataValidation(type="list", formula1=DECISIONS, allow_blank=True, showDropDown=False)
        dv.prompt = "Approved, Rejected, Watching or Needs Info"
        dv.promptTitle = "Reviewer decision"
        ws.add_data_validation(dv)
        dv.add(f"{decision_col}4:{decision_col}{3 + len(rows)}")
    else:
        ws.cell(row=4, column=1, value="No companies in this file.").font = BODY_FONT


def write_results(new_rows: list[dict[str, Any]], all_rows: list[dict[str, Any]], per_source: dict[str, int], counts: dict[str, int]) -> None:
    RESULTS.mkdir(parents=True, exist_ok=True)

    wb_master = Workbook()
    ws_master = wb_master.active
    ws_master.title = "Master"
    _write_sheet(ws_master, all_rows, f"All companies ever surfaced ({len(all_rows)}). Decisions persist across runs.")
    wb_master.save(MASTER)

    wb_latest = Workbook()
    ws_latest = wb_latest.active
    ws_latest.title = "Latest Run"
    source_line = ", ".join(f"{k}: {v}" for k, v in per_source.items())
    _write_sheet(
        ws_latest, new_rows,
        f"New and updated this run ({len(new_rows)}). Sources this run — {source_line}",
    )
    wb_latest.save(LATEST)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Run the multi-source prospecting pipeline")
    parser.add_argument("--sources", default="", help="Comma-separated source names (default: all)")
    parser.add_argument("--list-sources", action="store_true")
    parser.add_argument("--min-employees", type=int, default=500, help="Exclude companies smaller than this (default 500). Companies of unknown size are excluded, not assumed.")
    parser.add_argument("--max-employees", type=int, default=0, help="Exclude companies larger than this (0 = no cap, default)")
    parser.add_argument("--no-size-filter", action="store_true", help="Skip the employee-count lookup/filter entirely")
    parser.add_argument("--no-tier1a", action="store_true", help="Skip the company-site/jobs enrichment pass")
    parser.add_argument("--enrich-limit", type=int, default=40, help="Max near-miss companies to crawl for Tier 1A evidence (default 40)")
    args = parser.parse_args(argv)

    if args.list_sources:
        for name in intents_mod.SOURCES:
            print(name)
        return 0

    selected = [s.strip() for s in args.sources.split(",") if s.strip()] or None

    adopted = sync_decisions_from_master()
    if adopted:
        print(f"Read {adopted} reviewer decision(s) from last run's master.xlsx.\n")

    print(f"Running {len(selected or intents_mod.SOURCES)} source(s), {sum(intents_mod.summary().values())} intents total:\n")
    raw_items, per_source = run_sources(selected, args.min_employees, args.max_employees or None)

    sigs = build_signals(raw_items)
    print(f"\n{len(sigs)} unique signals after dedupe. Applying SOP gates...")

    records = [(sig, route(sig)) for sig in sigs]
    decisions = Counter(v.decision for _, v in records)
    excluded = Counter(v.exclusion_reason.split(":")[0] for _, v in records if v.exclusion_reason)

    companies = build_company_rows(records)

    sized_out = 0
    sizes: dict[str, int | None] = {}  # stays empty when sizing is skipped
    if args.no_size_filter:
        for c in companies:
            c["employees"] = ""
        # Without a size filter, carry forward the companies worth *investigating*,
        # not only those already passing. A company failing solely on "stage not
        # visible" or "no bottleneck" is precisely the one whose own pipeline and
        # careers pages can settle the question — enriching only already-passing
        # companies would waste the crawl on the cases that need it least.
        qualified = [c for c in companies if c["priority"] != "Reject"]
        near_miss = [
            c for c in companies
            if c["priority"] == "Reject" and (c.get("modality") or c.get("bottlenecks"))
        ]
        near_miss.sort(key=lambda c: -int(c.get("score") or 0))
        sized = qualified + near_miss[: args.enrich_limit]
        print(f"\n{len(qualified)} already qualified, {len(sized) - len(qualified)} "
              f"near-miss companies queued for Tier 1A evidence")
    else:
        # Size every non-excluded company, not just the ones that already scored
        # well — a real 90,000-person company mentioned in an SEC filing or a
        # ClinicalTrials.gov sponsor record deserves the same "known-large account"
        # floor account_based signals get, regardless of which of the 9 sources
        # found it. Restricting this to already-Review/A/B companies would miss
        # exactly the large accounts the small-company gates auto-reject.
        names = sorted({c["canonical_company"] for c in companies if c["canonical_company"]})
        print(f"\nLooking up employee counts for {len(names)} companies (cached lookups are free)...")
        sizes = company_size.enrich(names)

        max_emp = args.max_employees or None
        sized = []
        for c in companies:
            n = sizes.get(c["canonical_company"])
            c["employees"] = n
            if n is None or n < args.min_employees or (max_emp is not None and n > max_emp):
                if c["priority"] != "Reject":
                    sized_out += 1  # was otherwise qualified, just not the right size
                continue
            if c["priority"] == "Reject":
                # Confirmed large account, real on-topic signal, just couldn't clear
                # the small-company scoring bar (SOP: account-based only for this
                # population — a human researcher, not a formula, finds the named
                # discovery unit and buyer from here).
                #
                # The floor requires actual antibody/protein evidence. Without
                # this check it waved through any article that merely mentioned a
                # big company: a product-licensing story about "ProTek" surfaced
                # an unrelated 12,000-person firm called Protek as a prospect.
                if not (c.get("modality") or c.get("bottlenecks")):
                    continue
                c["priority"] = "Review"
                c["open_question"] = (
                    "Large account (SOP: account-based only) — identify the specific "
                    "discovery unit, program and buyer behind this signal before treating "
                    "it as qualified; a single signal cannot establish that on its own."
                )
            sized.append(c)

    # --- Tier 1A enrichment pass (brief §3) --------------------------------
    # Discovery sources find *who* to look at; the company's own site and job
    # board are the primary authority for stage and bottleneck, which is what
    # Gates 2/3/5 actually demand. Running this only on companies that already
    # survived sizing keeps the crawl small and targeted.
    if sized and not args.no_tier1a:
        names = [c["canonical_company"] for c in sized if c["canonical_company"]]
        print(f"\nTier 1A: resolving domains for {len(names)} companies...")
        profiles = company_profile.resolve(names)
        targets = [
            (c["canonical_company"], profiles[c["canonical_company"]].domain)
            for c in sized
            if profiles.get(c["canonical_company"]) and profiles[c["canonical_company"]].domain
        ]
        print(f"  {len(targets)}/{len(names)} have a resolvable website")

        for c in sized:
            profile = profiles.get(c["canonical_company"])
            if profile and profile.domain and not c.get("domain"):
                c["domain"] = profile.domain

        if targets:
            site_items = company_site.fetch(companies=targets)
            job_items = jobs.fetch(companies=targets)
            print(f"  company pages: {len(site_items)} usable | job postings: {len(job_items)}")
            # Clear the resolved domain before normalization. build_company_rows
            # keys a company on `domain or canonical_name`, so leaving it set
            # gave Tier 1A signals the key "aptevotherapeutics.com" while the
            # discovery signals for the same company keyed on "aptevo" — two
            # separate groups, and the merge below then discarded every page of
            # first-party evidence. The domain is still recorded on the company
            # row above; it must not be what the grouping hinges on here.
            for item in site_items:
                item["_source_name"] = "company_site"
                item["_known_domain"] = ""
            for item in job_items:
                item["_source_name"] = "jobs"
                item["_known_domain"] = ""
            per_source["company_site"] = len(site_items)
            per_source["jobs"] = len(job_items)

            tier1a = site_items + job_items
            if tier1a:
                extra_sigs = build_signals(tier1a)
                records.extend((sig, route(sig)) for sig in extra_sigs)
                # Rebuild company rows so the richer evidence actually moves
                # the score, then re-apply the size filter using cached sizes.
                rebuilt = build_company_rows(records)
                keep = {c["company_key"] for c in sized}
                merged = []
                for c in rebuilt:
                    if c["company_key"] not in keep:
                        continue
                    if not args.no_size_filter:
                        c["employees"] = sizes.get(c["canonical_company"])
                    # Promote only when first-party evidence actually arrived.
                    # Promoting on the pre-existing modality/bottleneck alone
                    # would mark every near-miss as Review whether or not the
                    # crawl found anything, which inflates the queue with the
                    # same records that already failed their gates.
                    got_first_party = any(
                        s in (c.get("sources") or "") for s in ("company_site", "jobs")
                    )
                    if c["priority"] == "Reject" and got_first_party and (
                        c.get("modality") or c.get("bottlenecks")
                    ):
                        c["priority"] = "Review"
                    merged.append(c)
                merged = [c for c in merged if c["priority"] != "Reject"]
                if merged:
                    sized = merged

    annotated, counts = registry.apply(sized)
    fresh = [r for r in annotated if r["movement"] in ("New", "Updated")]

    write_results(fresh, registry.all_rows(), per_source, counts)

    priorities = Counter(c["priority"] for c in companies)
    print(f"\nSignal decisions: {dict(decisions)}")
    if excluded:
        print(f"Hard exclusions:  {dict(excluded)}")
    print(f"Companies:        {len(companies)}  {dict(priorities)}")
    if not args.no_size_filter:
        cap_note = f"-{args.max_employees}" if args.max_employees else "+"
        print(f"Size filter:      {len(sized)} companies at {args.min_employees}{cap_note} employees "
              f"({sized_out} otherwise-qualified companies excluded for being outside that range)")
    print(f"Since last run:   {counts['New']} new, {counts['Updated']} updated, "
          f"{counts['Unchanged']} unchanged (skipped — already known), "
          f"{counts['Suppressed']} closed by reviewer (skipped)")
    print(f"\n>> {MASTER}")
    print(f">> {LATEST}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
