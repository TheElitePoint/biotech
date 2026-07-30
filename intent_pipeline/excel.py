"""Weekly Excel report.

Three sheets:

    New This Week   companies that are new or moved since the last run — the queue
    All Companies   everything ever surfaced, with reviewer decisions preserved
    Run Summary     counts, driven by formulas so it recalculates if rows are edited

The reviewer works in the Excel file. Their decisions are read back into
company_history.csv on the next run, which is what stops a rejected company
reappearing week after week.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPORT = Path(__file__).resolve().parent.parent / "output" / "prospect_review_queue.xlsx"

FONT = "Arial"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name=FONT, size=10)
LINK_FONT = Font(name=FONT, size=10, color="0563C1", underline="single")
# Yellow marks the cells the reviewer is expected to fill in.
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")

PRIORITY_FILL = {
    "A": PatternFill("solid", fgColor="C6EFCE"),
    "B": PatternFill("solid", fgColor="D9EAD3"),
    "Review": PatternFill("solid", fgColor="FFF2CC"),
}
MOVEMENT_FILL = {
    "New": PatternFill("solid", fgColor="D9E1F2"),
    "Updated": PatternFill("solid", fgColor="FCE4D6"),
}

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (key, header, width). Reviewer-owned columns come last so they are easy to reach.
COLUMNS = [
    ("movement", "Movement", 11),
    ("priority", "Priority", 9),
    ("score", "Score", 7),
    ("canonical_company", "Company", 26),
    ("domain", "Domain", 22),
    ("signal_types", "Intent Trigger", 16),
    ("top_signal_date", "Signal Date", 12),
    ("modality", "Modality", 18),
    ("asset_stage", "Asset Stage", 20),
    ("bottlenecks", "Bottleneck", 30),
    ("top_signal_title", "Evidence Headline", 52),
    ("top_signal_url", "Evidence URL", 42),
    ("open_question", "Blocking Question", 40),
    ("failed_gates", "Failed Gates", 34),
    ("signal_count", "Signals", 8),
    ("times_seen", "Runs Seen", 10),
    ("first_seen", "First Seen", 12),
    ("reviewer_status", "DECISION", 14),
    ("reviewer", "Reviewer", 14),
    ("review_date", "Review Date", 12),
    ("reviewer_notes", "Notes", 40),
]

REVIEWER_COLS = {"reviewer_status", "reviewer", "review_date", "reviewer_notes"}

DECISIONS = '"Approved,Rejected,Watching,Needs Info"'


def _write_sheet(ws, rows: list[dict[str, Any]], title_note: str) -> None:
    ws.sheet_view.showGridLines = False

    ws.cell(row=1, column=1, value=title_note).font = Font(
        name=FONT, size=11, bold=True, color="1F3864"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    for col, (_, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 28

    for i, row in enumerate(rows, start=4):
        for col, (key, _, _) in enumerate(COLUMNS, start=1):
            value = row.get(key, "")
            if key in ("score", "signal_count", "times_seen"):
                value = int(value) if str(value).strip().isdigit() else value
            cell = ws.cell(row=i, column=col, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=key in ("bottlenecks", "top_signal_title", "open_question", "failed_gates", "reviewer_notes"))

            if key == "top_signal_url" and value:
                cell.hyperlink = value
                cell.font = LINK_FONT
            elif key == "priority" and value in PRIORITY_FILL:
                cell.fill = PRIORITY_FILL[value]
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key == "movement" and value in MOVEMENT_FILL:
                cell.fill = MOVEMENT_FILL[value]
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key in REVIEWER_COLS:
                cell.fill = INPUT_FILL

        ws.row_dimensions[i].height = 30

    if rows:
        ws.freeze_panes = "D4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUMNS))}{3 + len(rows)}"

        decision_col = get_column_letter(
            next(i for i, (k, _, _) in enumerate(COLUMNS, start=1) if k == "reviewer_status")
        )
        dv = DataValidation(type="list", formula1=DECISIONS, allow_blank=True, showDropDown=False)
        dv.prompt = "Approved, Rejected, Watching or Needs Info"
        dv.promptTitle = "Reviewer decision"
        ws.add_data_validation(dv)
        dv.add(f"{decision_col}4:{decision_col}{3 + len(rows)}")
    else:
        ws.cell(row=4, column=1, value="No new or changed companies this run.").font = BODY_FONT


def _summary_sheet(ws, new_rows: int, all_rows: int, counts: dict[str, int]) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 52

    ws["A1"] = f"Prospect Review Queue — run {date.today().isoformat()}"
    ws["A1"].font = Font(name=FONT, size=13, bold=True, color="1F3864")

    ws["A3"] = "How to use this file"
    ws["A3"].font = Font(name=FONT, size=11, bold=True)
    guide = [
        "1. Work the 'New This Week' tab. It only holds companies that are new or have moved.",
        "2. For each row: open the Evidence URL, then the company's own pipeline page.",
        "3. Decide whether they own a therapeutic antibody or protein program.",
        "4. Set DECISION (yellow column). Approved / Rejected / Watching / Needs Info.",
        "5. Save the file. The next run reads your decisions and will not resurface closed rows.",
    ]
    for i, line in enumerate(guide, start=4):
        ws.cell(row=i, column=1, value=line).font = BODY_FONT
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)

    ws["A11"] = "This run"
    ws["A11"].font = Font(name=FONT, size=11, bold=True)

    # Data starts at row 4. With no rows, a range like R4:R3 is malformed, so the
    # count formulas fall back to literal zeros.
    last = 3 + all_rows
    has_rows = all_rows > 0

    def count(formula: str, fallback: int = 0) -> Any:
        return formula if has_rows else fallback

    metrics = [
        ("New companies", count('=COUNTIF(\'New This Week\'!A:A,"New")', new_rows), "First time ever surfaced"),
        ("Updated companies", count('=COUNTIF(\'New This Week\'!A:A,"Updated")'), "Seen before, new evidence this run"),
        ("Unchanged (not shown)", counts.get("Unchanged", 0), "In history, nothing new — omitted"),
        ("Closed by reviewer (suppressed)", counts.get("Suppressed", 0), "Already decided; will not resurface"),
        ("Priority A", count(f"=COUNTIF('All Companies'!B4:B{last},\"A\")"), "Score 80+"),
        ("Priority B", count(f"=COUNTIF('All Companies'!B4:B{last},\"B\")"), "Score 68-79"),
        ("Awaiting decision", count(f"=COUNTBLANK('All Companies'!R4:R{last})"), "DECISION column still blank"),
        ("Total companies tracked", all_rows, "All time"),
    ]
    for i, (label, value, note) in enumerate(metrics, start=12):
        ws.cell(row=i, column=1, value=label).font = BODY_FONT
        cell = ws.cell(row=i, column=2, value=value)
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=3, value=note).font = Font(name=FONT, size=9, color="595959")

    ws["A22"] = "Scoring thresholds (SOP p24): 80+ = A, 68-79 = B, 55-67 = Review, below 55 = Reject."
    ws["A22"].font = Font(name=FONT, size=9, italic=True, color="595959")
    ws.merge_cells("A22:C22")
    ws["A23"] = "No company reaches Approve automatically. A human confirms asset ownership first (SOP Gate 5)."
    ws["A23"].font = Font(name=FONT, size=9, italic=True, color="595959")
    ws.merge_cells("A23:C23")


def sync_decisions() -> int:
    """Read the reviewer's decisions out of last run's workbook into history.

    Runs before the new report is written, since writing replaces the file. If the
    reviewer has the file open or it is missing, this is a no-op rather than an error —
    losing a run's classification is recoverable, crashing the pipeline is worse.
    """
    if not REPORT.exists():
        return 0

    try:
        from openpyxl import load_workbook

        wb = load_workbook(REPORT, data_only=True)
    except Exception:
        return 0

    keys = [key for key, _, _ in COLUMNS]
    decisions: dict[str, dict[str, str]] = {}

    for sheet in ("All Companies", "New This Week"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or not any(row):
                continue
            record = {k: ("" if v is None else str(v)) for k, v in zip(keys, row)}
            company = record.get("canonical_company", "").strip()
            domain = record.get("domain", "").strip()
            key = domain or company.lower()
            status = record.get("reviewer_status", "").strip()
            if key and status:
                decisions[key] = record

    from . import history

    return history.record_decisions(decisions)


def write(new_rows: list[dict[str, Any]], all_rows: list[dict[str, Any]], counts: dict[str, int]) -> Path:
    wb = Workbook()

    summary = wb.active
    summary.title = "Run Summary"

    ws_new = wb.create_sheet("New This Week")
    _write_sheet(
        ws_new,
        new_rows,
        f"New and changed companies — {date.today().isoformat()} "
        f"({len(new_rows)} to review). Fill the yellow DECISION column.",
    )

    ws_all = wb.create_sheet("All Companies")
    _write_sheet(
        ws_all,
        all_rows,
        f"Every company surfaced to date ({len(all_rows)}). Decisions here persist across runs.",
    )

    _summary_sheet(summary, len(new_rows), len(all_rows), counts)

    REPORT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(REPORT)
    return REPORT
