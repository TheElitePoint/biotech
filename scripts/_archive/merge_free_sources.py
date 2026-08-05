"""Merge all free, live data sources into one deduplicated sheet.

Sources combined (all re-run live, nothing recycled):
  - output/universe_intent.csv     ClinicalTrials.gov + NIH RePORTER (via build_final.py)
  - output/daily_company_candidates.csv    Bing News RSS
  - output/daily_publication_candidates.csv Europe PMC

No date columns are included in the output, per request.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

OUT = Path(__file__).resolve().parent / "output"
MERGED_CSV = OUT / "merged_all_sources.csv"
MERGED_XLSX = OUT / "merged_all_sources.xlsx"


def norm_key(name: str) -> str:
    name = re.sub(
        r"[,\s]+(inc|llc|ltd|limited|corp|corporation|co|plc|gmbh|ag|ab|sa|nv|bv|"
        r"pte|pty|holdings?|group)\.?$",
        "",
        name.strip(),
        flags=re.I,
    )
    return re.sub(r"[^a-z0-9]", "", name.lower())


def read(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))


def from_universe_intent() -> list[dict[str, str]]:
    rows = read(OUT / "universe_intent.csv")
    out = []
    label = {
        "registry": "ClinicalTrials.gov",
        "reporter": "NIH RePORTER",
        "reporter+registry": "NIH RePORTER + ClinicalTrials.gov",
    }
    for r in rows:
        out.append(
            {
                "Company": r.get("canonical_company", ""),
                "Priority": r.get("priority", ""),
                "Score": r.get("score", ""),
                "Source": label.get(r.get("trigger", ""), r.get("trigger", "")),
                "Company Website": "",
                "Modality": r.get("modality", ""),
                "Program / Intervention": r.get("interventions", "")[:300],
                "Bottleneck": r.get("bottlenecks", ""),
                "Evidence Summary": r.get("evidence", ""),
                "Evidence URL": r.get("evidence_url", ""),
                "Note": r.get("note", ""),
                "_key": r.get("company_key") or norm_key(r.get("canonical_company", "")),
            }
        )
    return out


def from_daily(path_name: str, source_label: str) -> list[dict[str, str]]:
    rows = read(OUT / path_name)
    out = []
    for r in rows:
        name = r.get("Current Company Name", "")
        if not name:
            continue
        out.append(
            {
                "Company": name,
                "Priority": r.get("Corrected Status", ""),
                "Score": r.get("Total Score", ""),
                "Source": source_label,
                "Company Website": r.get("Company Website", ""),
                "Modality": r.get("Confirmed Modality", ""),
                "Program / Intervention": r.get("Therapeutic Asset or Program", ""),
                "Bottleneck": r.get("Scientific / Development Requirement", ""),
                "Evidence Summary": r.get("Signal Summary", ""),
                "Evidence URL": r.get("Original Trigger Source URL", ""),
                "Note": r.get("Research Notes", ""),
                "_key": r.get("_company_key") or norm_key(name),
            }
        )
    return out


def merge() -> list[dict[str, str]]:
    all_rows = (
        from_universe_intent()
        + from_daily("daily_company_candidates.csv", "Bing News RSS")
        + from_daily("daily_publication_candidates.csv", "Europe PMC")
    )

    def score_of(row: dict[str, str]) -> float:
        try:
            return float(row.get("Score") or 0)
        except ValueError:
            return 0.0

    by_key: dict[str, dict[str, str]] = {}
    for row in all_rows:
        key = row["_key"]
        if not key:
            continue
        if key not in by_key:
            by_key[key] = row
            continue
        existing = by_key[key]
        if row["Source"] not in existing["Source"]:
            existing["Source"] = f"{existing['Source']}; {row['Source']}"
        if score_of(row) > score_of(existing):
            for field in ("Priority", "Score", "Modality", "Program / Intervention",
                          "Bottleneck", "Evidence Summary", "Evidence URL"):
                if row.get(field):
                    existing[field] = row[field]
        if row.get("Company Website") and not existing.get("Company Website"):
            existing["Company Website"] = row["Company Website"]

    merged = list(by_key.values())
    merged.sort(key=lambda r: -score_of(r))
    return merged


def write_csv(rows: list[dict[str, str]]) -> None:
    fields = ["Company", "Priority", "Score", "Source", "Company Website", "Modality",
              "Program / Intervention", "Bottleneck", "Evidence Summary", "Evidence URL", "Note"]
    with MERGED_CSV.open("w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def write_xlsx(rows: list[dict[str, str]]) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    fields = ["Company", "Priority", "Score", "Source", "Company Website", "Modality",
              "Program / Intervention", "Bottleneck", "Evidence Summary", "Evidence URL", "Note"]
    widths = [26, 9, 7, 30, 22, 26, 34, 30, 50, 42, 40]
    priority_fill = {
        "A": PatternFill("solid", fgColor="C6EFCE"),
        "B": PatternFill("solid", fgColor="D9EAD3"),
        "Review": PatternFill("solid", fgColor="FFF2CC"),
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Sources Merged"
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        ws.append([row.get(f, "") for f in fields])

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        priority_val = r[1].value
        if priority_val in priority_fill:
            r[1].fill = priority_fill[priority_val]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{1 + len(rows)}"
    wb.save(MERGED_XLSX)


def main() -> int:
    rows = merge()
    write_csv(rows)
    write_xlsx(rows)
    from collections import Counter
    print(f"{len(rows)} unique companies merged -> {MERGED_CSV} / {MERGED_XLSX}")
    print(f"  priority: {dict(Counter(r['Priority'] for r in rows))}")
    src_counter: Counter = Counter()
    for r in rows:
        for s in r["Source"].split("; "):
            src_counter[s] += 1
    print(f"  by source: {dict(src_counter)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
